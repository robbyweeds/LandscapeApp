from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers, colors
import datetime
import sqlite3



def createWorkbook(db):

    wb = Workbook()

    ws = wb.active
    ws.column_dimensions['A'].width= 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    ws['A1'] = datetime.date.today()

    thick_border = Side(border_style="thick", color="000000")
    thin_border = Side(border_style="thin", color="000000")

    ws['A7'] = 'Notes:'

    ws['B7']= 'qty'
    ws['B7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['B7'].alignment = Alignment(horizontal='center')
    ws['B7'].font = Font(bold=True, size= 12)
    ws['C7'] = 'descriptions'
    ws['C7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['C7'].alignment = Alignment(horizontal='center')
    ws['C7'].font = Font(bold=True, size= 12)
    ws['D7'] = 'unit'
    ws['D7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['D7'].alignment = Alignment(horizontal='center')
    ws['D7'].font = Font(bold=True, size= 12)
    ws['E7'] = 'unit cost'
    ws['E7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['E7'].alignment = Alignment(horizontal='center')
    ws['E7'].font = Font(bold=True, size= 12)
    ws['F7'] = 'ext cost'
    ws['F7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['F7'].alignment = Alignment(horizontal='center')
    ws['F7'].font = Font(bold=True, size= 12)
    ws['G7'] = 'labor factor'
    ws['G7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['G7'].alignment = Alignment(horizontal='center')
    ws['G7'].font = Font(bold=True, size= 12)
    ws['H7'] = 'man hours'
    ws['H7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['H7'].alignment = Alignment(horizontal='center')
    ws['H7'].font = Font(bold=True, size= 12)

    conn = sqlite3.connect(db)

    cur = conn.cursor()

    plant_data = cur.execute('''SELECT * FROM plants''').fetchall()

    plantrows = len(plant_data)

    for i in plant_data:
        this_row = str(plant_data.index(i) + 8)
        qty_col = 'B' + this_row
        desc_col = 'C' +this_row
        unit_col = 'D' +this_row
        unit_cost_col = 'E' +this_row
        ext_cost_col = 'F' +this_row
        labor_factor_col = 'G' +this_row
        manhour_col = 'H' +this_row
        ws[qty_col] = float(i[1])
        ws[qty_col].alignment = Alignment(horizontal='center')
        ws[qty_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[desc_col] = i[0]
        ws[desc_col].alignment = Alignment(horizontal='center')
        ws[desc_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[unit_col] = i[4]
        ws[unit_col].alignment = Alignment(horizontal='center')
        ws[unit_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[unit_cost_col] = float(i[3])
        ws[unit_cost_col].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws[unit_cost_col].alignment = Alignment(horizontal='center')
        ws[unit_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[ext_cost_col] = float(i[1]) * float(i[3])
        ws[ext_cost_col].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws[ext_cost_col].alignment = Alignment(horizontal='center')
        ws[ext_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[labor_factor_col] = 0
        ws[labor_factor_col].alignment = Alignment(horizontal='center')
        ws[labor_factor_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[manhour_col] = 0
        ws[manhour_col].alignment = Alignment(horizontal='center')
        ws[manhour_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

    
    direct_row = 'C' + str(plantrows+ 10)
    ws[direct_row] = 'DIRECT COST LABOR'
    ws[direct_row].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thin_border)
    ws[direct_row].font = Font(bold=False, size= 9)
    ws[direct_row].alignment = Alignment(horizontal='center')
    direct_mat_row = 'C' + str(plantrows+11)
    ws[direct_mat_row] = 'DIRECT COST MATERIALS(Materials, Tax, Freight)'
    ws[direct_mat_row].border = Border(top=thin_border, left=thick_border, right=thick_border, bottom=thin_border)
    ws[direct_mat_row].font = Font(bold=False, size= 8)
    ws[direct_mat_row].alignment = Alignment(horizontal='center')
    billable_eqip_row = 'C' + str(plantrows+12)
    ws[billable_eqip_row] = 'Billable Equipment Rate'
    ws[billable_eqip_row].border = Border(top=thin_border, left=thick_border, right=thick_border, bottom=thin_border)
    ws[billable_eqip_row].font = Font(bold=False, size= 9)
    ws[billable_eqip_row].alignment = Alignment(horizontal='center')
    total_direct_row = 'C' + str(plantrows+13)
    ws[total_direct_row] = 'TOTAL DIRECT COST'
    ws[total_direct_row].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws[total_direct_row].font = Font(bold=False, size= 9)
    ws[total_direct_row].alignment = Alignment(horizontal='center')
    desired_markup_row = 'C' + str(plantrows+15)
    ws[desired_markup_row] = 'Enter Desired Mat Markup %'
    ws[desired_markup_row].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws[desired_markup_row].font = Font(bold=True, size= 9)
    ws[desired_markup_row].alignment = Alignment(horizontal='center')
    ws[desired_markup_row].fill = PatternFill('solid', start_color="ffff00")



    wb.save("workbook.xlsx")

